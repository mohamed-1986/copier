#! Python 3
# - Copy and Paste Ranges using OpenPyXl library
import openpyxl ,datetime, os, xlrd, re
# excel loading
# file name is given, the output is sheets to copy from
import time


def TheSheet97(file):
    if file.endswith(".xls"):
        wb= xlrd.open_workbook(file)
        sheets = [ws for ws in wb.sheet_names() if ws =="EMC" or "AREA" in ws.upper()]
        return wb, sheets
    
def Area(Tag):
    unit= re.findall('\d+', Tag)[0]
    if unit in ['51','52','53','54','55','56','57','58']:
        return "A4/"+ unit
    if unit in ['02','03','04','05','06','12','16','39']:
        return "2A/"+ unit
    if unit in ['07','13','14','15','45']:
        return "2B/"+ unit
    else:
        return unit


#Copy range of cells as a nested list.[{"Tag":"02-FT-010","Problem":"Blockage"},{"Tag":"04-FV-002","Problem":"Stuck"}]
def copyRange(copyDict, sheet):
    startRow = searchRowStarting(sheet, "TAG")[0]
    rangeSelected = []
    #Loops through selected Rows. the while is to loop until data is finished
    try:
        while '-' in str( sheet.cell(startRow, copyDict["Tag"]).value) or 'UNIT' in str(sheet.cell(startRow, copyDict["Tag"]).value).upper():
            rowSelected = {}
            for j in copyDict:
                try:
                    rowSelected[j] =str(sheet.cell(startRow, copyDict[j]).value)
            #Adds the RowSelected List and nests inside the rangeSelected   
                except :
                    print("wrong index {}".format(j))
            rangeSelected.append(rowSelected)
            startRow= startRow + 1
    except IndexError:
        pass
    return rangeSelected

#Paste data from copyRange into template sheet
def pasteRange(copyDict, pasteDict, sheetReceiving, copiedData, datePaste):
    countRow = 0
    startRow= sheetReceiving.max_row +1
    # #Check last row that it is not empty
    while str(sheetReceiving.cell(startRow-1, pasteDict["Tag"] ).value)== "None":
        startRow= startRow-1         #decrement the row to check if the data was manualy deleted

    endRow= startRow+ len(copiedData)
    # pasting is in here: #first we set the serial then the date
    for i in range(startRow, endRow,1):  # for every row, we start to paste the new row
        if isinstance(sheetReceiving.cell(i-1,1).value,int)== True:
            try:  #pasting a new serial number
                sheetReceiving.cell(i,1).value = sheetReceiving.cell(i-1,1).value +1
            except:
                sheetReceiving.cell(i,1).value= 1
        sheetReceiving.cell(i,pasteDict["Date"]).value = datePaste    # pasting date
        #  Second we paste corrsponding data: tag of copy with tag of paste, problem with problem, etc..
        for j in copyDict:   #for every column, we paste matched tags
            sheetReceiving.cell(i,pasteDict[j] ).value = copiedData[countRow][j]
        try:
            sheetReceiving.cell(i,pasteDict["Loc"]).value = Area(sheetReceiving.cell(i,pasteDict["Tag"]).value)
        except:
            pass
        countRow += 1
#this returns the column place for the main headers ex. tag in column 2 and problem in column 3
def searchRowStarting(sheet, theWord):
    for i in range(1, 7,1):
        #Appends the row to a RowSelected list
        for j in range(1, 7,1):
            try:
                if theWord in str(sheet.cell(i,j).value).upper():
                    if '-' in str(sheet.cell(i+1,j).value):
                        return i+1, j, "first", sheet.cell(i+1,j).value
                    elif '-' in str(sheet.cell(i+2,j).value):
                        return i+2, j, "second", sheet.cell(i+2,j).value 
                    elif 'UNIT' in str(sheet.cell(i+1,j).value).upper():
                        return i+1, j, "first", sheet.cell(i+1,j).value
                    elif 'UNIT' in str(sheet.cell(i+2 ,j).value).upper():
                        return i+2, j, "first", sheet.cell(i+2 ,j).value
                    else:
                        return FileExistsError
            except IndexError:
                return FileExistsError

def searchForWord(sheet, theWord):
    for i in range(1,sheet.nrows):
        if i> 7:
            break
        else:
            for j in range(1,sheet.ncols):
                if j>7:
                    break
                elif theWord in str(sheet.cell(i,j).value).upper():
                    return i, j

def searchForWordXlsx(sheet, theWord):
    for i in range(1,12):
        for j in range(1,12):
            if theWord in str(sheet.cell(i,j).value).upper():
                return i, j

# def moveData(copyFileName, copyFileSheet, pasteFileName, pasteFileSheet):
def moveData97(copyFileName, copyFileSheet, pasteFileName):
    # t0= time.perf_counter()
    if copyFileName.endswith(".xls"):
        wb = xlrd.open_workbook(copyFileName)
        copySheet= wb.sheet_by_name(copyFileSheet)
    # t1= time.perf_counter()
    pasteFile = openpyxl.load_workbook(pasteFileName)
    pasteSheet= pasteFile.active

    pasteDict, copyDict, dateCopy= dictionaryxls( copyFileName, copySheet, pasteSheet)
    try:
        if searchRowStarting(copySheet,"TAG") != FileExistsError and searchRowStarting(copySheet,"TAG") != IndexError:
            selectedRange = copyRange(copyDict, copySheet)
            pasteRange(copyDict, pasteDict, pasteSheet, selectedRange, dateCopy) 
            pasteFile.save(pasteFileName)
            return True
    except:
        return False

def dictionaryxls(copyFileName, copySheet, pasteSheet ):
    copyDict={}
    pasteDict= {}
    olddateCopy = os.path.basename(copyFileName)
    if  len(olddateCopy) < 16:
        dateCopy = olddateCopy.split('.')[0]
    else :
        dateCopy= (olddateCopy.split('-')[-3])[-2] + (olddateCopy.split('-')[-3])[-1] +'-' 
        dateCopy= dateCopy+ (olddateCopy.split('-')[-2]+'-'+ olddateCopy.split('-')[-1]).split('.')[-2]
    dateCopy= dateCopy.replace('-','/')

    try:
        pasteDict["Tag"]= searchForWordXlsx(pasteSheet, "TAG")[1]
    except TypeError:
        pass
    try:
        pasteDict["Loc"]= searchForWordXlsx(pasteSheet, "LOC")[1]
    except TypeError:
        pass
    try:
        copyDict["Tag"]= searchForWord(copySheet, "TAG")[1]
    except TypeError:
        pass
    try:
        pasteDict["Problem"]= searchForWordXlsx(pasteSheet,"PROB")[1]
    except TypeError:
        pass
    try:
        copyDict["Problem"]= searchForWord(copySheet,"PROB")[1]
    except TypeError:
        pass
    try:
        pasteDict["Complain"]= searchForWordXlsx(pasteSheet,"COMPLAIN")[1]
    except TypeError:
        pass
    try:
        copyDict["Complain"]= searchForWord(copySheet,"COMPLAIN")[1]
    except TypeError:
        pass
    try:
        pasteDict["Action"]= searchForWordXlsx(pasteSheet,"ACTION")[1]
    except TypeError:
        pass
    try:
        copyDict["Action"]= searchForWord(copySheet,"ACTION")[1]
    except TypeError:
        pass
    try:
        pasteDict["Date"]= searchForWordXlsx(pasteSheet,"DATE")[1]
    except TypeError:
        pass
    return pasteDict, copyDict, dateCopy